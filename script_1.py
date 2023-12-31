import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.pyplot import figure

import matplotlib.gridspec as gridspec
import itertools

# Set the initial value for excel_file
excel_file_path = ""
sheet_names = []
start = 0
end = 0
increment = 0

options = [
    "Bench"
]

columnOptions = ["N/A"]
global selected_sheet
selected_sheet = ""
filterIndices = []
masterDF = pd.DataFrame()

class InvalidFilterInputError(Exception):
    pass

def validate_input(user_input):
    valid_chars = set("0123456789-,[]")
    return all(char in valid_chars for char in user_input)
    #returns true if all the characters are valid. Returns false if there are funky characters

# def parse_filter_ranges(filter_ranges):
#     parsed_ranges = []
#     for filter_range in filter_ranges:
#         filter_range = filter_range.strip()  # Remove any leading/trailing spaces
#         try:
#             if "-" in filter_range:
#                 # Parse the value as a range
#                 start, end = map(int, filter_range.split("-"))
#                 parsed_ranges.append(list(range(start, end + 1)))
#             elif filter_range.startswith("[") and filter_range.endswith("]"):
#                 # Parse the value as a list
#                 list_values = filter_range[1:-1].split(",")
#                 filter_values = [int(value.strip()) for value in list_values]
#                 parsed_ranges.append(filter_values)
#             else:
#                 # Parse the value as a single integer
#                 parsed_ranges.append([int(filter_range)])
#         except ValueError:
#             messagebox.showerror("Invalid Input", "Invalid filter value. Please enter a valid integer, range, or list.")
#     return parsed_ranges

def parse_filter_ranges(input_str):
    def expand_range(range_str):
        start, end = map(int, range_str.split("-"))
        return list(range(start, end + 1))
    
    masterDF = pd.read_excel(excel_file_path, sheet_name= selected_sheet, engine="openpyxl")
    filterIndices = filter_listbox.curselection()
    index = 0
    result = []
    listEntryColumnCheck = []
    elements = input_str.split(", ")

    for element in elements:
        if element.startswith("[") and element.endswith("]"):
            try:
                result.append([int(x) for x in element[1:-1].split(",")])
            except ValueError:
                listEntryColumnCheck.extend([x] for x in element[1:-1].split(","))

                for item in listEntryColumnCheck:
                    if item in masterDF[columnOptions[filterIndices[index]]].values:
                        result.append(item)
                    else:
                        messagebox.showerror("Invalid Input", "One of the filter values you provided is not found in the excel spreadsheet")

        elif "-" in element:
            try:
                result.append(expand_range(element))
            except ValueError:
                messagebox.showerror("Invalid Input", "Invalid filter range. You may have made a typo. Please enter a valid range")

        else:
            try:
                elem = int(element)
                result.append(([elem]))
            except ValueError:
                if element in masterDF[columnOptions[filterIndices[index]]].values:
                    result.append([element])
                else:
                    messagebox.showerror("Invalid Input", "One of the filter values you provided is not found in the excel spreadsheet")


        index += 1

    #     final_result = []
    #     temp_list = []

    #     for item in result:
    #         if isinstance(item, list):
    #             if temp_list:
    #                 final_result.append(temp_list)
    #                 temp_list = []
    #             final_result.append(item)
    #         else:
    #             temp_list.append(int(item))

    #     if temp_list:
    #         final_result.append(temp_list)
    # except ValueError:
    #     messagebox.showerror("Invalid Input", "Invalid filter value. You may have made a typo. Please enter a valid integer, range, or list.")

    return result



def openFile():
    global excel_file_path
    file_paths = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx; *.xls"), ("CSV Files", "*.csv")])
    if file_paths:
        file_path = file_paths[0]
        excel_file_path = file_path
        sheet_names = update_sheet_names()
        update_file_label(excel_file_path)
        update_sheet_box(sheet_names)

def update_file_label(file_path):
    file_label.config(text=file_path)

def get_selected_sheet():
    global selected_sheet
    global columnOptions
    selected_indices = sheetBox.curselection()
    if selected_indices:
        selected_index = selected_indices[0]
        selected_sheet = sheetBox.get(selected_index)
        print("Selected sheet:", selected_sheet)
        columnOptions = get_sheet_columns(selected_sheet)
        # Update the dropdown menus
        update_dropdown_menus()
    else:
        messagebox.showwarning("Warning", "No sheet selected!")

def update_dropdown_menus():
    x_dropdown['menu'].delete(0, 'end')
    y_dropdown['menu'].delete(0, 'end')
    filter_listbox.delete(0, 'end')

    for option in columnOptions:
        x_dropdown['menu'].add_command(label=option, command=tk._setit(x_var, option))
        y_dropdown['menu'].add_command(label=option, command=tk._setit(y_var, option))
        filter_listbox.insert(tk.END, option)

def update_sheet_box(sheet_names):
    sheetBox.delete(0, tk.END)
    for sheet in sheet_names:
        sheetBox.insert(tk.END, sheet)

def get_sheet_columns(sheet_name):
    if excel_file_path and sheet_name:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[sheet_name]
        columnOptions1 = [cell.value for cell in sheet[1]]
        workbook.close()
        return columnOptions1
    return columnOptions1

def setHardcodedPath():
    global excel_file_path
    excel_file_path = "/Users/brindha/Downloads/movandi2 thing.xlsx"
    sheet_names = update_sheet_names()
    update_file_label(excel_file_path)
    update_sheet_box(sheet_names)

def update_sheet_names():
    if excel_file_path:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet_names = workbook.sheetnames
        workbook.close()
        return sheet_names
    return []

def spacing_list(start, end, increment):
    integer_list = []
    current_value = start

    if(start > end): 
        currrent_value = start
        while (current_value >= end):
            integer_list.append(current_value)
            current_value -= increment
        return integer_list

    elif (start < end):
        current_value = start
        while (current_value <= end):
                integer_list.append(current_value)
                current_value += increment
        return integer_list
    else:
        # If the increment is zero, return an empty list
        return []

    return integer_list


def string_to_int(intList):
    string_list = []

    # Iterate through each integer in the input list
    for num in intList:
        string_list.append(str(num))

    return string_list

# def on_motion(event):
#     # Update the x= and y= values in the plot's title
#     if event.xdata is not None and event.ydata is not None:
#         plt.gca().set_title(f"x={event.xdata:.2f}, y={event.ydata:.2f}")



    

def plotGraph():
    #plt.subplots_adjust(left=0.15)
    #fig = plt.figure()
    #left_subplot = fig.add_subplot(121)
    #ax.set_box_aspect(1)
    # ax.set_adjustable('box')
    # plt.rcParams['figure.figsize'] = [8, 6]
    # fig.set_figwidth(8)
    # fig.set_figheight(6)  
    figure(figsize=(8,8), dpi=80)


    filterNamesList = []
    combinationList = []
    legendLabels = []
    xPoints = []
    yPoints = []

    masterDF = pd.read_excel(excel_file_path, sheet_name= selected_sheet, engine="openpyxl")
    filterIndices = filter_listbox.curselection()

    if(len(filterIndices) != 0):
        if(filter_range_entry.get().__eq__("")):
            messagebox.showerror("Error", "Please enter at least one filter range that corresponds to the selected filter name")
            return

    if(len(filterIndices) == 0):
        messagebox.showerror("Error", "Please select at least 1 filter name")
        return

    # Retrieve the selected x-axis, y-axis, filter indices, and filter ranges
    xName = x_var.get()
    yName = y_var.get()
    filterRangesList = parse_filter_ranges(filter_range_entry.get())

    # Check if all the required fields are selected
    if(not (xName and yName)):
        messagebox.showerror("Error", "Please select both an X-axis name and a Y-axis name")
        return

    if len(filterIndices) != len(filterRangesList):
        # Display an error message if the number of filter names and ranges don't match
        messagebox.showerror("Error", "Number of filter names and ranges should match.")
        return

    if(xName and yName and filterIndices and filterRangesList):

        # Read the data from the Excel file
        tempMasterDF = masterDF
        filterNamesList = [columnOptions[idx] for idx in filter_listbox.curselection()]

        '''for filter_name, filter_range in zip(filterNamesList, filterRangesList):
            masterDF = masterDF[masterDF[filter_name].isin(filter_range)]'''
        
        for combination in itertools.product(*filterRangesList):
            combinationList.append(combination)
            legendLabels.append(combination)
            print(combination)

        for currentCombo, label in zip(combinationList, legendLabels):
            tempMasterDF = masterDF
            for currentFilterNum, currentFilterName in zip(currentCombo, filterNamesList):
                tempMasterDF = tempMasterDF[tempMasterDF[currentFilterName] == currentFilterNum]
            xPoints = tempMasterDF[xName]
            yPoints = tempMasterDF[yName]
            plt.grid(True)
            plt.plot(xPoints, yPoints, marker='o', label = label)

        legendTitle = ",".join(filterNamesList)


        plot_data = []
        legend_labels = []

        # Set the x-axis, y-axis labels, title, and formatting options
        
        plt.ylabel(yName)
        plt.xlabel(xName)
        plt.title("Graph of " + xName + " vs " + yName)
        plt.xticks(rotation='vertical')

        legend = plt.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop = {'size': 10})
        plt.tight_layout()
        plt.subplots_adjust(left=0.05, right=0.7)
        legend.set_title(legendTitle)
        



        # Set the scrollable window to the canvas size
        scroll_window.configure(scrollregion=scroll_window.bbox(tk.ALL))

        if (y_spacing_entry.get()):
            try:
                y_spacing = int(y_spacing_entry.get())
            except (ValueError, InvalidFilterInputError):
                messagebox.showerror("Invalid Input", "Invalid entry for the 'Y-axis step'. Please enter a valid integer.")
                plt.ioff()
                plt.close()
                return


        if (y_start_entry.get() and y_end_entry.get()):

            try:
                y_start = int(y_start_entry.get())
            except (ValueError, InvalidFilterInputError):
                messagebox.showerror("Invalid Input", "Invalid entry for the 'Y-axis Starting Value'. Please enter a valid integer.")
                plt.ioff()
                plt.close()
                return

            try:
                y_end = int(y_end_entry.get())
            except (ValueError, InvalidFilterInputError):
                messagebox.showerror("Invalid Input", "Invalid entry for the 'Y-axis Ending Value'. Please enter a valid integer.")
                plt.ioff()
                plt.close()
                return
            
            plt.autoscale(False, tight=False)

            if(y_start > y_end):
                plt.ylim([y_end, y_start])
            else:
                plt.ylim([y_start, y_end])

            if(y_spacing_entry.get()):
                y_tick_list = spacing_list(y_start, y_end, y_spacing)
                y_tick_list.sort(reverse = False)
                y_tick_label = string_to_int(y_tick_list)
                plt.yticks(y_tick_list, y_tick_label)
            else:
                y_spacing = abs((y_start - y_end)/10)
                y_tick_list = spacing_list(y_start, y_end, y_spacing)
                y_tick_list.sort(reverse = False)
                y_tick_label = string_to_int(y_tick_list)
                plt.yticks(y_tick_list, y_tick_label)


        elif (not y_start_entry.get() and y_end_entry.get()):
            messagebox.showerror("Error", "If you fill in an ending value for the Y-axis, you must fill a starting value. Another option is to leave both parameters blank and the graph will automatically scale the Y-axis. Please try again")
            plt.ioff()  # Turn off interactive mode to avoid multiple plot windows
            plt.close()  # Close the previous plot window
            return  # This will exit the function and prevent the graph plot from opening

        elif (not y_end_entry.get() and y_start_entry.get()):
            messagebox.showerror("Error", "If you fill in a starting value for the Y-axis, you must fill an ending value. Another option is to leave both parameters blank and the graph will automatically scale the Y-axis. Please try again")
            plt.ioff()  # Turn off interactive mode to avoid multiple plot windows
            plt.close()  # Close the previous plot window
            return


        if (x_start_entry.get() and x_end_entry.get()):

            try:
                x_start = int(x_start_entry.get())
            except (ValueError, InvalidFilterInputError):
                messagebox.showerror("Invalid Input", "Invalid entry for the 'X-axis Starting Value'. Please enter a valid integer.")
                plt.ioff()
                plt.close()
                return   
            
            try:
                x_end = int(x_end_entry.get())
            except (ValueError, InvalidFilterInputError):
                messagebox.showerror("Invalid Input", "Invalid entry for the 'X-axis Ending Value'. Please enter a valid integer.")
                plt.ioff()
                plt.close()
                return

            plt.autoscale(False, tight=False)

            if(x_start > x_end):
                plt.xlim([x_end, x_start])
            else:
                plt.xlim([x_start, x_end])

            if(x_spacing_entry.get()):
                try:
                    x_spacing = int(x_spacing_entry.get())
                except (ValueError, InvalidFilterInputError):
                    messagebox.showerror("Invalid Input", "Invalid entry for the 'X-axis Step'. Please enter a valid integer.")
                    plt.ioff()
                    plt.close()
                    return               

                x_tick_list = (spacing_list(x_start, x_end, x_spacing))
                x_tick_list.sort(reverse = False)
                x_tick_labels = string_to_int(x_tick_list)
                plt.xticks(x_tick_list, x_tick_labels)
            else:
                x_spacing = abs((x_start - x_end)/10)
                x_tick_list = spacing_list(x_start, x_end, x_spacing)
                x_tick_list.sort(reverse = False)
                x_tick_labels = string_to_int(x_tick_list)
                plt.xticks(x_tick_list, x_tick_labels)               

        elif (not x_start_entry.get() and x_end_entry.get()):
            messagebox.showerror("Error", "If you fill in an ending value for the X-axis, you must fill a starting value. Another option is to leave both parameters blank and the graph will automatically scale the X-axis. Please try again")
            plt.ioff()  # Turn off interactive mode to avoid multiple plot windows
            plt.close()  # Close the previous plot window
            return
        elif (not x_end_entry.get() and x_start_entry.get()):
            messagebox.showerror("Error", "If you fill in a starting value for the X-axis, you must fill an ending value. Another option is to leave both parameters blank and the graph will automatically scale the X-axis. Please try again")
            plt.ioff()  # Turn off interactive mode to avoid multiple plot windows
            plt.close()  # Close the previous plot window
            return



        # Display the plot
        plt.ion()
        #plt.connect('motion_notify_event', on_motion)
        plt.show(block = False)
        


    else:
        # Display an error message if any of the required fields are not selected
        messagebox.showerror("Error", "Please select all fields.")




# Create the GUI
window = tk.Tk()
window.title("Graph GUI")

# Create a scrollbar
scrollbar = ttk.Scrollbar(window)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Create a scrollable window
scroll_window = tk.Canvas(window, yscrollcommand=scrollbar.set)
scroll_window.pack(fill=tk.BOTH, expand=True)

# Configure the scrollbar to work with the scrollable window
scrollbar.config(command=scroll_window.yview)

# Create the main frame
frame = tk.Frame(scroll_window)
scroll_window.create_window((0, 0), window=frame, anchor=tk.NW)

# Create the button to open the file
open_file_button = tk.Button(frame, text="Open File", command=openFile)
open_file_button.pack()

# Create the button to set the hardcoded path
hardcoded_file_button = tk.Button(frame, text="Hardcoded Data File (brindha)", command=setHardcodedPath)
hardcoded_file_button.pack()

sheetName = tk.Label(frame, text="Select Sheet Name:")
sheetName.pack()
sheetBox = tk.Listbox(frame, selectmode=tk.SINGLE)
for sheet in sheet_names:
    sheetBox.insert(tk.END, sheet)
sheetBox.pack()

processSheet = tk.Button(frame, text="Process Selected Sheet", command=get_selected_sheet)
processSheet.pack()

file_label = tk.Label(frame, text="No file selected")
file_label.pack()

# Create the dropdown menus for x-axis, y-axis, and filter options
x_label = tk.Label(frame, text="xName:")
x_label.pack()
x_var = tk.StringVar(frame)
x_dropdown = tk.OptionMenu(frame, x_var, *columnOptions)
x_dropdown.pack()

y_label = tk.Label(frame, text="yName:")
y_label.pack()
y_var = tk.StringVar(frame)
y_dropdown = tk.OptionMenu(frame, y_var, *columnOptions)
y_dropdown.pack()

filter_label = tk.Label(frame, text="filterName:")
filter_label.pack()
filter_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE)
for option in columnOptions:
    filter_listbox.insert(tk.END, option)
filter_listbox.pack()

filter_range_label = tk.Label(frame, text="filterRanges (ex: 1-40, b1_1h, 25, [1,4,7], ['b1_1h, b1_2h'])")
filter_range_label.pack()
filter_range_entry = tk.Entry(frame)
filter_range_entry.pack()

#Customize spacing
x_start_label = tk.Label(frame, text = "X-axis Starting Value")
x_start_label.pack()
x_start_entry = tk.Entry(frame)
x_start_entry.pack()

x_end_label = tk.Label(frame, text = "X-axis Ending Value")
x_end_label.pack()
x_end_entry = tk.Entry(frame)
x_end_entry.pack()

x_spacing_label = tk.Label(frame, text = "X-axis step")
x_spacing_label.pack()
x_spacing_entry = tk.Entry(frame, text = "X-axis step")
x_spacing_entry.pack()

y_start_label = tk.Label(frame, text = "Y-axis Starting Value")
y_start_label.pack()
y_start_entry = tk.Entry(frame)
y_start_entry.pack()

y_end_label = tk.Label(frame, text = "Y-axis Ending Value")
y_end_label.pack()
y_end_entry = tk.Entry(frame)
y_end_entry.pack()

y_spacing_label = tk.Label(frame, text = "Y-axis step")
y_spacing_label.pack()
y_spacing_entry = tk.Entry(frame, text = "Y-axis step")
y_spacing_entry.pack()


# Create the button to plot the graph
plot_button = tk.Button(frame, text="Plot Graph", command=plotGraph)
plot_button.pack()

# Start the GUI event loop
window.mainloop()


